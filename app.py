from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st


APP_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = APP_DIR / "property_listings_MM.xlsx"
REVIEWS_PATH = APP_DIR / "unit_viewing_reviews.csv"
URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)
NORMALIZED_TEXT_PATTERN = re.compile(r"[^a-z0-9]+")
MRT_MINUTES_PATTERN = re.compile(r"\((\d+)\s*mins?\s*\)", re.IGNORECASE)
MRT_ENTRY_PATTERN = re.compile(
    r"(?:[A-Z]{1,3}\d+[A-Z]?(?:/[A-Z]{1,3}\d+[A-Z]?)*)\s+([A-Za-z][A-Za-z'&./\- ]+?)\s*\(\s*(\d+)\s*mins?\s*\)",
    re.IGNORECASE,
)

YES_NO_OPTIONS = ["Unknown", "Yes", "No", "Need to verify"]
UTILITIES_OPTIONS = [
    "Unknown",
    "Yes - utilities and air-con included",
    "No",
    "Partially",
    "Need to verify",
]
SUN_OPTIONS = ["Unknown", "Sunrise", "Sunset", "Both", "Need to verify"]
KITCHEN_OPTIONS = ["Not applicable", "Common kitchen", "Private kitchen", "Need to verify"]
VIEWING_OUTCOME_OPTIONS = ["Pending", "OK", "Not OK"]
SINGAPORE_TIMEZONE = ZoneInfo("Asia/Singapore")
GOOGLE_CALENDAR_GUESTS = [
    "abdumuhaiminhashim@gmail.com",
    "maisarahkm@gmail.com",
]
BOOKING_TIME_PATTERN = re.compile(
    r"^\s*(\d{1,2})\s+([A-Za-z]+),\s*[A-Za-z]{3}\s*@\s*(\d{1,2}:\d{2}\s*[ap]m)\s*$",
    re.IGNORECASE,
)

CHECKLIST_FIELDS = [
    ("utilities_aircon_inclusive", "Price inclusive of utilities + air-con?"),
    ("water_heater", "Water heater"),
    ("washing_machine", "Washing machine"),
    ("fridge", "Fridge"),
    ("bedroom_ceiling_fan", "Ceiling fan in bedroom"),
    ("living_room_ceiling_fan", "Ceiling fan in living room"),
    ("floor_level", "Floor level of unit"),
    ("sun_exposure", "Unit gets sunrise or sunset?"),
    ("resident_profile", "Group of people that stays around the building"),
    ("co_living_kitchen", "Co-living kitchen in common area or not?"),
]

REVIEW_COLUMNS = [
    "listing_key",
    "viewed_unit",
    "viewing_outcome",
    "utilities_aircon_inclusive",
    "water_heater",
    "washing_machine",
    "fridge",
    "bedroom_ceiling_fan",
    "living_room_ceiling_fan",
    "floor_level",
    "sun_exposure",
    "resident_profile",
    "co_living_kitchen",
    "overall_notes",
    "last_updated",
]


def text_value(value: object, fallback: str = "") -> str:
    if value is None:
        return fallback
    if isinstance(value, float) and pd.isna(value):
        return fallback
    text = str(value).strip()
    return text or fallback


def numeric_value(value: object) -> float | None:
    series = pd.to_numeric(pd.Series([value]), errors="coerce")
    number = series.iloc[0]
    if pd.isna(number):
        return None
    return float(number)


def unique_preserving_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def slugify(text: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return cleaned or "listing"


def parse_note_fields(row: pd.Series, note_columns: list[str]) -> tuple[str, str]:
    notes: list[str] = []
    listing_url = ""
    for column in note_columns:
        value = text_value(row.get(column))
        if not value:
            continue
        urls = URL_PATTERN.findall(value)
        if urls and not listing_url:
            listing_url = urls[0].rstrip(")")
        cleaned = URL_PATTERN.sub("", value).strip(" |,;")
        if cleaned:
            notes.append(cleaned)
    return " | ".join(unique_preserving_order(notes)), listing_url


def normalize_search_text(text: str) -> str:
    return NORMALIZED_TEXT_PATTERN.sub(" ", text.lower()).strip()


def clean_mrt_station_name(segment: str) -> str:
    without_codes = re.sub(
        r"^(?:[A-Z]{1,3}\d+[A-Z]?(?:/[A-Z]{1,3}\d+[A-Z]?)*)\s+",
        "",
        segment.strip(),
    )
    without_minutes = re.sub(r"\(\s*\d+\s*mins?\s*\)", "", without_codes, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", without_minutes).strip()


def parse_mrt_entries(value: object) -> list[tuple[str, int | None]]:
    raw_value = text_value(value)
    if not raw_value:
        return []

    entries: list[tuple[str, int | None]] = []
    matched_entries = MRT_ENTRY_PATTERN.findall(raw_value)
    if matched_entries:
        for station_name, minutes in matched_entries:
            cleaned_station = re.sub(r"\s+", " ", station_name).strip()
            entries.append((cleaned_station, int(minutes)))
    else:
        segments = re.split(r"\s{2,}|\|", raw_value)
        for segment in segments:
            cleaned_station = clean_mrt_station_name(segment)
            if not cleaned_station:
                continue
            minute_match = MRT_MINUTES_PATTERN.search(segment)
            minutes = int(minute_match.group(1)) if minute_match else None
            entries.append((cleaned_station, minutes))

    best_by_station: dict[str, int | None] = {}
    station_order: list[str] = []
    for station, minutes in entries:
        if station not in best_by_station:
            best_by_station[station] = minutes
            station_order.append(station)
            continue
        current = best_by_station[station]
        if current is None or (minutes is not None and minutes < current):
            best_by_station[station] = minutes

    return [(station, best_by_station[station]) for station in station_order]


def parse_mrt_stations(value: object) -> list[str]:
    return [station for station, _ in parse_mrt_entries(value)]


def parse_mrt_minutes(value: object) -> list[int]:
    minutes = [minutes for _, minutes in parse_mrt_entries(value) if minutes is not None]
    if minutes:
        return minutes

    raw_value = text_value(value)
    if not raw_value:
        return []
    return [int(match) for match in MRT_MINUTES_PATTERN.findall(raw_value)]


def best_mrt_minutes(value: object) -> int | None:
    minutes = parse_mrt_minutes(value)
    return min(minutes) if minutes else None


def nearest_mrt_station(value: object) -> str:
    entries = [(station, minutes) for station, minutes in parse_mrt_entries(value) if minutes is not None]
    if entries:
        return min(entries, key=lambda entry: entry[1])[0]

    stations = parse_mrt_stations(value)
    return stations[0] if stations else ""


def build_search_blob(
    listing_name: str,
    address: str,
    nearest_mrt: str,
    notes: str,
    mrt_stations: list[str],
    booking_time: str,
    agent_contact: str,
) -> str:
    parts = [
        listing_name,
        address,
        nearest_mrt,
        notes,
        " ".join(mrt_stations),
        booking_time,
        agent_contact,
    ]
    if nearest_mrt:
        parts.append("mrt")
    return " ".join(normalize_search_text(part) for part in parts if part)


def build_mrt_search_blob(nearest_mrt: str, mrt_stations: list[str]) -> str:
    parts = [nearest_mrt, " ".join(mrt_stations)]
    if nearest_mrt or mrt_stations:
        parts.append("mrt")
    return " ".join(normalize_search_text(part) for part in parts if part)


def search_matches_query(search_blob: str, mrt_search_blob: str, query: str) -> bool:
    normalized_query = normalize_search_text(query)
    if not normalized_query:
        return True

    query_tokens = normalized_query.split()
    if query_tokens[0] == "mrt":
        if len(query_tokens) == 1:
            return bool(mrt_search_blob)
        return bool(mrt_search_blob) and all(token in mrt_search_blob for token in query_tokens[1:])

    return all(token in search_blob for token in query_tokens)


def is_co_living_listing(text: str) -> bool:
    lowered = text.lower()
    return any(keyword in lowered for keyword in ["co-living", "co living", "coliwoo"])


def build_group_key(row: dict[str, object]) -> str:
    listing_url = text_value(row.get("listing_url"))
    if listing_url:
        return listing_url.lower()
    address = text_value(row.get("address"))
    listing_name = text_value(row.get("listing_name"))
    return slugify(f"{listing_name}-{address}")


def format_currency(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "Not listed"
    return f"S${value:,.0f}"


def format_number(value: float | None, suffix: str = "") -> str:
    if value is None or pd.isna(value):
        return "Not listed"
    if float(value).is_integer():
        return f"{int(value):,}{suffix}"
    return f"{value:,.2f}{suffix}"


def listing_label(row: pd.Series) -> str:
    outcome = text_value(row.get("viewing_outcome"))
    status = text_value(row.get("review_status"))
    if outcome == "OK":
        badge = "✓"
    elif outcome == "Not OK":
        badge = "✗"
    elif status == "Complete":
        badge = "●"
    elif status == "In progress":
        badge = "◐"
    else:
        badge = "○"
    booking_confirmed = text_value(row.get("booking_confirmed"))
    if booking_confirmed == "confirmed":
        booking_suffix = " · ✓ confirmed"
    elif booking_confirmed == "unconfirmed":
        booking_suffix = " · ⚠ unconfirmed"
    else:
        booking_suffix = ""
    return f"{badge} {row['listing_name']} | {format_currency(row['monthly_rent'])} | {row['source_sheet']}{booking_suffix}"


def _booking_font_status(cell: object) -> str | None:
    """Return 'confirmed' for red font, 'unconfirmed' for orange font, None otherwise."""
    try:
        color = cell.font.color  # type: ignore[attr-defined]
        if color.type == "rgb":
            rgb: str = color.rgb  # e.g. "FFFF0000" (alpha + R + G + B)
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
            if r > 150 and g < 100 and b < 100:
                return "confirmed"
            if r > 150 and 80 <= g < 200 and b < 80:
                return "unconfirmed"
    except Exception:
        pass
    return None


def extract_booking_confirmation(workbook_source: str | BytesIO) -> dict[tuple[str, int], str]:
    import openpyxl
    wb = openpyxl.load_workbook(workbook_source, data_only=True)
    result: dict[tuple[str, int], str] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        booking_col: int | None = None
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value and "booking" in str(cell.value).lower():
                booking_col = cell.column
                break
        if booking_col is None:
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            booking_cell = row[booking_col - 1]
            if booking_cell.value:
                status = _booking_font_status(booking_cell)
                if status is not None:
                    result[(sheet_name, row_idx)] = status
    return result


def parse_booking_datetime(value: object, now: datetime | None = None) -> datetime | None:
    booking_time = text_value(value)
    if not booking_time:
        return None

    match = BOOKING_TIME_PATTERN.match(booking_time)
    if not match:
        return None

    day, month_name, time_text = match.groups()
    reference = now or datetime.now(SINGAPORE_TIMEZONE)
    parsed = datetime.strptime(
        f"{int(day):02d} {month_name} {reference.year} {time_text.upper().replace(' ', '')}",
        "%d %B %Y %I:%M%p",
    ).replace(tzinfo=SINGAPORE_TIMEZONE)

    if parsed < reference - timedelta(days=180):
        parsed = parsed.replace(year=parsed.year + 1)

    return parsed


def escape_ics_text(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\n", "\\n")
    )


def build_calendar_event_details(selected: pd.Series) -> tuple[datetime, datetime, str, list[str]] | None:
    start_time = parse_booking_datetime(selected.get("booking_time"))
    if start_time is None:
        return None

    end_time = start_time + timedelta(hours=1)
    address = text_value(selected.get("address"))
    location = address if address and address != "Address not provided" else text_value(selected.get("listing_name"))
    description_parts = [
        f"Listing: {text_value(selected.get('listing_name'))}",
        f"Booking time: {text_value(selected.get('booking_time'))}",
    ]
    if address and address != "Address not provided":
        description_parts.append(f"Address: {address}")
    agent_contact = text_value(selected.get("agent_contact"))
    if agent_contact:
        description_parts.append(f"Agent / Contact: {agent_contact}")
    notes = text_value(selected.get("notes"))
    if notes and notes != "No notes in source sheet":
        description_parts.append(f"Notes: {notes}")
    listing_url = text_value(selected.get("listing_url"))
    if listing_url:
        description_parts.append(f"Listing URL: {listing_url}")

    return start_time, end_time, location, description_parts


def build_google_calendar_url(selected: pd.Series) -> str | None:
    event_details = build_calendar_event_details(selected)
    if event_details is None:
        return None

    start_time, end_time, location, description_parts = event_details
    params = {
        "action": "TEMPLATE",
        "text": f"Unit Viewing - {text_value(selected.get('listing_name'))}",
        "dates": f"{start_time.strftime('%Y%m%dT%H%M%S')}/{end_time.strftime('%Y%m%dT%H%M%S')}",
        "ctz": "Asia/Singapore",
        "details": "\n".join(description_parts),
        "location": location,
        "add": ",".join(GOOGLE_CALENDAR_GUESTS),
    }
    return "https://calendar.google.com/calendar/render?" + urlencode(params)


def build_calendar_invite(selected: pd.Series) -> tuple[str, str] | None:
    event_details = build_calendar_event_details(selected)
    if event_details is None:
        return None

    start_time, end_time, location, description_parts = event_details

    now_utc = datetime.now(ZoneInfo("UTC")).strftime("%Y%m%dT%H%M%SZ")
    uid = f"{slugify(text_value(selected.get('listing_key')))}-{start_time.strftime('%Y%m%dT%H%M%S')}@unit-viewing"
    ics_body = "\r\n".join(
        [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//Unit Viewing Dashboard//EN",
            "CALSCALE:GREGORIAN",
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_utc}",
            f"DTSTART;TZID=Asia/Singapore:{start_time.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND;TZID=Asia/Singapore:{end_time.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{escape_ics_text('Unit Viewing - ' + text_value(selected.get('listing_name')))}",
            f"LOCATION:{escape_ics_text(location)}",
            f"DESCRIPTION:{escape_ics_text(chr(10).join(description_parts))}",
            *[
                f"ATTENDEE;ROLE=REQ-PARTICIPANT;PARTSTAT=NEEDS-ACTION;RSVP=TRUE:mailto:{email}"
                for email in GOOGLE_CALENDAR_GUESTS
            ],
            "END:VEVENT",
            "END:VCALENDAR",
            "",
        ]
    )
    file_name = f"{slugify(text_value(selected.get('listing_name')))}-unit-viewing.ics"
    return ics_body, file_name


def review_defaults() -> dict[str, str]:
    return {
        "listing_key": "",
        "viewed_unit": "No",
        "viewing_outcome": VIEWING_OUTCOME_OPTIONS[0],
        "utilities_aircon_inclusive": UTILITIES_OPTIONS[0],
        "water_heater": YES_NO_OPTIONS[0],
        "washing_machine": YES_NO_OPTIONS[0],
        "fridge": YES_NO_OPTIONS[0],
        "bedroom_ceiling_fan": YES_NO_OPTIONS[0],
        "living_room_ceiling_fan": YES_NO_OPTIONS[0],
        "floor_level": "",
        "sun_exposure": SUN_OPTIONS[0],
        "resident_profile": "",
        "co_living_kitchen": KITCHEN_OPTIONS[0],
        "overall_notes": "",
        "last_updated": "",
    }


def workbook_signature(workbook_path: Path) -> tuple[int, int]:
    stat = workbook_path.stat()
    return stat.st_mtime_ns, stat.st_size


def build_listing_records(
    workbook: pd.ExcelFile,
    booking_confirmation: dict[tuple[str, int], str] | None = None,
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for sheet_name in workbook.sheet_names:
        frame = pd.read_excel(workbook, sheet_name=sheet_name)
        note_columns = [column for column in ["Notes", "Notes.1", "Unnamed: 12"] if column in frame.columns]
        for position, (index, row) in enumerate(frame.iterrows()):
            notes, listing_url = parse_note_fields(row, note_columns)
            listing_name = text_value(row.get("Listing Name"), fallback=f"Listing {position + 1}")
            address = text_value(row.get("Address"), fallback="Address not provided")
            nearest_mrt = text_value(row.get("Nearest MRT station"))
            booking_time = text_value(row.get("Booking time"))
            agent_contact = text_value(row.get("Agent name / Contact No"))
            booking_confirmed = (booking_confirmation or {}).get((sheet_name, position), "" if not booking_time else "confirmed")
            mrt_stations = parse_mrt_stations(nearest_mrt)
            nearest_mrt_minutes = best_mrt_minutes(nearest_mrt)
            nearest_mrt_station_name = nearest_mrt_station(nearest_mrt)
            description_text = " ".join(
                [
                    listing_name,
                    address,
                    text_value(row.get("Property Type")),
                    notes,
                    listing_url,
                ]
            )
            record = {
                "listing_key": f"{sheet_name}:{text_value(row.get('#'), str(position + 1))}:{slugify(listing_name)}",
                "listing_group": "",
                "source_sheet": sheet_name,
                "listing_number": text_value(row.get("#"), str(position + 1)),
                "listing_name": listing_name,
                "address": address,
                "nearest_mrt": nearest_mrt,
                "mrt_stations": mrt_stations,
                "mrt_station_names": ", ".join(mrt_stations),
                "nearest_mrt_station": nearest_mrt_station_name,
                "nearest_mrt_minutes": nearest_mrt_minutes,
                "monthly_rent": numeric_value(row.get("Monthly Rent (S$)")),
                "beds": text_value(row.get("Beds"), "Not listed"),
                "baths": text_value(row.get("Baths"), "Not listed"),
                "size_sqft": numeric_value(row.get("Size (sqft)")),
                "psf_sgd": numeric_value(row.get("PSF (S$)")),
                "property_type": text_value(row.get("Property Type"), "Unknown"),
                "furnishing": text_value(row.get("Furnishing"), "Unknown"),
                "lease_period": text_value(row.get("Lease Period"), "Not listed"),
                "year_built": text_value(row.get("Year Built"), "Not listed"),
                "booking_time": booking_time,
                "booking_confirmed": booking_confirmed,
                "agent_contact": agent_contact,
                "notes": notes or "No notes in source sheet",
                "listing_url": listing_url,
                "search_blob": build_search_blob(
                    listing_name,
                    address,
                    nearest_mrt,
                    notes,
                    mrt_stations,
                    booking_time,
                    agent_contact,
                ),
                "mrt_search_blob": build_mrt_search_blob(nearest_mrt, mrt_stations),
                "is_co_living": is_co_living_listing(description_text),
            }
            record["listing_group"] = build_group_key(record)
            records.append(record)
    return pd.DataFrame(records)


@st.cache_data(show_spinner=False)
def load_listings(workbook_path: str, _workbook_signature: tuple[int, int]) -> pd.DataFrame:
    booking_confirmation = extract_booking_confirmation(workbook_path)
    workbook = pd.ExcelFile(workbook_path)
    return build_listing_records(workbook, booking_confirmation)


@st.cache_data(show_spinner=False)
def load_uploaded_listings(workbook_bytes: bytes, workbook_name: str) -> pd.DataFrame:
    booking_confirmation = extract_booking_confirmation(BytesIO(workbook_bytes))
    workbook = pd.ExcelFile(BytesIO(workbook_bytes))
    return build_listing_records(workbook, booking_confirmation)


@st.cache_data(show_spinner=False)
def load_reviews(review_path: str) -> pd.DataFrame:
    if not Path(review_path).exists():
        return pd.DataFrame(columns=REVIEW_COLUMNS)
    reviews = pd.read_csv(review_path, keep_default_na=False)
    for column in REVIEW_COLUMNS:
        if column not in reviews.columns:
            reviews[column] = ""
    return reviews[REVIEW_COLUMNS].fillna("")


def save_review(record: dict[str, str]) -> None:
    existing = load_reviews(str(REVIEWS_PATH)).copy()
    incoming = pd.DataFrame([record], columns=REVIEW_COLUMNS)
    if existing.empty:
        updated = incoming
    else:
        updated = pd.concat(
            [existing[existing["listing_key"] != record["listing_key"]], incoming],
            ignore_index=True,
        )
    updated = updated.sort_values(by="last_updated", ascending=False)
    updated.to_csv(REVIEWS_PATH, index=False)
    load_reviews.clear()


def field_answered(key: str, value: object, is_co_living: bool) -> bool:
    text = text_value(value)
    if key == "co_living_kitchen":
        if not is_co_living:
            return False
        return text not in {"", "Unknown", "Not applicable", "Need to verify"}
    if key in {"floor_level", "resident_profile"}:
        return bool(text)
    return text not in {"", "Unknown", "Need to verify"}


def applicable_checklist_keys(is_co_living: bool) -> list[str]:
    keys = [field_key for field_key, _ in CHECKLIST_FIELDS if field_key != "co_living_kitchen"]
    if is_co_living:
        keys.append("co_living_kitchen")
    return keys


def has_viewing_activity(row: pd.Series) -> bool:
    viewed_unit = text_value(row.get("viewed_unit"))
    viewing_outcome = text_value(row.get("viewing_outcome"))
    return viewed_unit == "Yes" or viewing_outcome in {"OK", "Not OK"}


def add_review_summary(data: pd.DataFrame) -> pd.DataFrame:
    data = data.copy()
    answered_counts: list[int] = []
    total_counts: list[int] = []
    completion_labels: list[str] = []
    for _, row in data.iterrows():
        checklist_keys = applicable_checklist_keys(bool(row["is_co_living"]))
        total_fields = len(checklist_keys)
        answered_fields = sum(
            field_answered(field_key, row.get(field_key), bool(row["is_co_living"]))
            for field_key in checklist_keys
        )
        review_started = answered_fields > 0 or has_viewing_activity(row)
        answered_counts.append(answered_fields)
        total_counts.append(total_fields)
        if not review_started:
            completion_labels.append("Not reviewed")
        elif answered_fields >= total_fields:
            completion_labels.append("Complete")
        else:
            completion_labels.append("In progress")
    data["answered_fields"] = answered_counts
    data["total_fields"] = total_counts
    data["completion_ratio"] = data["answered_fields"] / data["total_fields"]
    data["review_status"] = completion_labels
    return data


def build_dashboard_data(listings: pd.DataFrame, reviews: pd.DataFrame) -> pd.DataFrame:
    dashboard = listings.merge(reviews, on="listing_key", how="left")
    defaults = review_defaults()
    for column, default in defaults.items():
        if column == "listing_key":
            continue
        dashboard[column] = dashboard[column].fillna(default)
    dashboard = add_review_summary(dashboard)
    # Deduplicate across sheets: per listing_group keep the most-reviewed copy,
    # falling back to original sheet order when review data is equal.
    dashboard = (
        dashboard
        .sort_values("answered_fields", ascending=False, kind="stable")
        .drop_duplicates(subset=["listing_group"], keep="first")
        .reset_index(drop=True)
    )
    return dashboard


def option_index(options: list[str], current: str) -> int:
    return options.index(current) if current in options else 0


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,700&family=Manrope:wght@400;500;700&display=swap');

        .stApp {
            background:
                radial-gradient(circle at top left, rgba(224, 242, 254, 0.9), transparent 28%),
                radial-gradient(circle at bottom right, rgba(252, 211, 77, 0.28), transparent 32%),
                linear-gradient(180deg, #f8fafc 0%, #fff7ed 100%);
            color: #1c1917;
        }

        .block-container {
            max-width: 1400px;
            padding-top: 2.25rem;
            padding-bottom: 3rem;
        }

        h1, h2, h3 {
            font-family: 'Fraunces', serif;
            letter-spacing: -0.03em;
        }

        p, span, label, div {
            font-family: 'Manrope', sans-serif;
        }

        .hero-shell {
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 28px;
            padding: 1.9rem 2rem;
            background: linear-gradient(135deg, rgba(15, 118, 110, 0.92), rgba(14, 116, 144, 0.88));
            color: #f8fafc;
            box-shadow: 0 18px 50px rgba(15, 23, 42, 0.18);
            margin-bottom: 1rem;
        }

        .hero-shell h1 {
            color: #f8fafc;
            margin-bottom: 0.4rem;
            font-size: 2.5rem;
        }

        .hero-shell p {
            color: rgba(248, 250, 252, 0.9);
            margin: 0;
            font-size: 1rem;
        }

        .surface-card {
            border: 1px solid rgba(15, 23, 42, 0.08);
            border-radius: 24px;
            padding: 1.2rem 1.3rem;
            background: rgba(255, 255, 255, 0.78);
            backdrop-filter: blur(14px);
            box-shadow: 0 10px 30px rgba(15, 23, 42, 0.08);
        }

        .stat-label {
            font-size: 0.8rem;
            text-transform: uppercase;
            letter-spacing: 0.12em;
            color: #155e75;
            margin-bottom: 0.2rem;
        }

        .stat-value {
            font-family: 'Fraunces', serif;
            font-size: 2rem;
            color: #082f49;
        }

        .detail-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 0.9rem;
            margin-top: 0.8rem;
        }

        .detail-tile {
            background: rgba(255, 255, 255, 0.62);
            border: 1px solid rgba(14, 116, 144, 0.14);
            border-radius: 18px;
            padding: 0.9rem;
        }

        .detail-tile .label {
            font-size: 0.75rem;
            text-transform: uppercase;
            letter-spacing: 0.1em;
            color: #0f766e;
            margin-bottom: 0.3rem;
        }

        .detail-tile .value {
            font-size: 1rem;
            color: #0f172a;
            font-weight: 700;
        }

        .pill {
            display: inline-block;
            padding: 0.35rem 0.75rem;
            border-radius: 999px;
            background: rgba(8, 47, 73, 0.08);
            color: #0f172a;
            font-size: 0.85rem;
            margin-right: 0.4rem;
            margin-bottom: 0.4rem;
        }

        .missing-list {
            border-left: 4px solid #fb923c;
            padding-left: 0.9rem;
            margin-top: 0.9rem;
            color: #7c2d12;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def filter_dashboard(data: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.header("Filters")
    search_text = st.sidebar.text_input("Search listing, address, MRT station, or notes")
    source_options = sorted(data["source_sheet"].dropna().unique().tolist())
    selected_sources = source_options.copy()
    if len(source_options) > 1:
        selected_sources = st.sidebar.multiselect(
            "Source lists",
            source_options,
            default=source_options,
        )

    mrt_station_options = sorted(
        {
            station
            for stations in data["mrt_stations"].dropna().tolist()
            for station in stations
            if station
        }
    )
    selected_mrt_stations = st.sidebar.multiselect("MRT stations", mrt_station_options)

    mrt_minutes_series = data["nearest_mrt_minutes"].dropna()
    selected_mrt_minutes: tuple[int, int] | None = None
    if not mrt_minutes_series.empty:
        mrt_minutes_min = int(mrt_minutes_series.min())
        mrt_minutes_max = int(mrt_minutes_series.max())
        selected_mrt_minutes = st.sidebar.slider(
            "Distance to MRT (mins)",
            min_value=mrt_minutes_min,
            max_value=mrt_minutes_max,
            value=(mrt_minutes_min, mrt_minutes_max),
            step=1,
        )

    property_options = sorted(data["property_type"].dropna().unique().tolist())
    selected_property_types = st.sidebar.multiselect(
        "Property type",
        property_options,
        default=property_options,
    )

    furnishing_options = sorted(data["furnishing"].dropna().unique().tolist())
    selected_furnishing = st.sidebar.multiselect(
        "Furnishing",
        furnishing_options,
        default=furnishing_options,
    )

    rent_series = data["monthly_rent"].dropna()
    rent_min = int(rent_series.min()) if not rent_series.empty else 0
    rent_max = int(rent_series.max()) if not rent_series.empty else 4000
    rent_range = st.sidebar.slider(
        "Monthly rent (S$)",
        min_value=rent_min,
        max_value=rent_max,
        value=(rent_min, rent_max),
        step=50,
    )

    co_living_filter = st.sidebar.selectbox(
        "Co-living",
        ["All listings", "Co-living only", "Exclude co-living"],
    )

    review_filter = st.sidebar.selectbox(
        "Review status",
        ["All", "Not reviewed", "In progress", "Complete"],
    )

    viewed_filter = st.sidebar.selectbox(
        "Viewed status",
        ["All", "Viewed only", "Not viewed"],
    )

    selected_outcomes = st.sidebar.multiselect(
        "Viewing outcome",
        VIEWING_OUTCOME_OPTIONS,
        default=VIEWING_OUTCOME_OPTIONS,
    )

    filtered = data.copy()
    if selected_sources:
        filtered = filtered[filtered["source_sheet"].isin(selected_sources)]
    if selected_property_types:
        filtered = filtered[filtered["property_type"].isin(selected_property_types)]
    if selected_furnishing:
        filtered = filtered[filtered["furnishing"].isin(selected_furnishing)]
    filtered = filtered[
        filtered["monthly_rent"].fillna(rent_min).between(rent_range[0], rent_range[1])
    ]

    if co_living_filter == "Co-living only":
        filtered = filtered[filtered["is_co_living"]]
    elif co_living_filter == "Exclude co-living":
        filtered = filtered[~filtered["is_co_living"]]

    if review_filter != "All":
        filtered = filtered[filtered["review_status"] == review_filter]

    if viewed_filter == "Viewed only":
        filtered = filtered[filtered["viewed_unit"] == "Yes"]
    elif viewed_filter == "Not viewed":
        filtered = filtered[filtered["viewed_unit"] != "Yes"]

    if selected_outcomes:
        filtered = filtered[filtered["viewing_outcome"].isin(selected_outcomes)]

    if selected_mrt_stations:
        filtered = filtered[
            filtered["mrt_stations"].apply(
                lambda stations: any(station in stations for station in selected_mrt_stations)
            )
        ]

    if selected_mrt_minutes is not None:
        mrt_distance_mask = filtered["nearest_mrt_minutes"].between(
            selected_mrt_minutes[0],
            selected_mrt_minutes[1],
        )
        filtered = filtered[mrt_distance_mask]

    if search_text:
        filtered = filtered[
            filtered.apply(
                lambda row: search_matches_query(row["search_blob"], row["mrt_search_blob"], search_text),
                axis=1,
            )
        ]

    filtered = filtered.sort_values(by=["monthly_rent", "listing_name"], na_position="last")
    st.sidebar.caption(f"Showing {len(filtered)} of {len(data)} listings")
    return filtered


def missing_checklist_items(row: pd.Series) -> list[str]:
    missing: list[str] = []
    for field_key, label in CHECKLIST_FIELDS:
        if field_key == "co_living_kitchen" and not bool(row["is_co_living"]):
            continue
        if not field_answered(field_key, row.get(field_key), bool(row["is_co_living"])):
            missing.append(label)
    return missing


def review_for_listing(listing_key: str, reviews: pd.DataFrame) -> dict[str, str]:
    defaults = review_defaults()
    if reviews.empty or listing_key not in reviews["listing_key"].values:
        defaults["listing_key"] = listing_key
        return defaults
    record = reviews.loc[reviews["listing_key"] == listing_key].iloc[0].to_dict()
    result: dict[str, str] = {}
    for column, default in defaults.items():
        result[column] = text_value(record.get(column), default)
    return result


def render_hero(filtered: pd.DataFrame) -> None:
    avg_rent = filtered["monthly_rent"].mean() if not filtered.empty else None
    viewed_units = int((filtered["viewed_unit"] == "Yes").sum())
    ok_outcomes = int((filtered["viewing_outcome"] == "OK").sum())
    st.markdown(
        f"""
        <div class="hero-shell">
            <h1>Unit Viewing Dashboard</h1>
            <p>Excel-backed review board for shortlist triage, viewing prep, and live checklist tracking.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    stats = st.columns(4)
    stat_values = [
        ("Visible entries", str(len(filtered))),
        ("Viewed units", str(viewed_units)),
        ("OK outcomes", str(ok_outcomes)),
        ("Average rent", format_currency(avg_rent)),
    ]
    for column, (label, value) in zip(stats, stat_values):
        with column:
            st.markdown(
                f"""
                <div class="surface-card">
                    <div class="stat-label">{label}</div>
                    <div class="stat-value">{value}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    now = datetime.now(SINGAPORE_TIMEZONE)
    upcoming_confirmed = []
    upcoming_unconfirmed = []
    for _, row in filtered.iterrows():
        dt = parse_booking_datetime(row.get("booking_time"), now)
        if dt and now <= dt <= now + timedelta(days=7):
            confirmed = text_value(row.get("booking_confirmed"))
            entry = (dt, text_value(row.get("listing_name")))
            if confirmed == "confirmed":
                upcoming_confirmed.append(entry)
            else:
                upcoming_unconfirmed.append(entry)
    if upcoming_confirmed or upcoming_unconfirmed:
        upcoming_confirmed.sort()
        upcoming_unconfirmed.sort()
        parts = []
        if upcoming_confirmed:
            confirmed_lines = [
                f"• **{name}** — {dt.strftime('%a, %d %b @ %I:%M %p')} ✓"
                for dt, name in upcoming_confirmed
            ]
            parts.append("**✓ Confirmed**\n\n" + "\n\n".join(confirmed_lines))
        if upcoming_unconfirmed:
            unconfirmed_lines = [
                f"• **{name}** — {dt.strftime('%a, %d %b @ %I:%M %p')} ⚠"
                for dt, name in upcoming_unconfirmed
            ]
            parts.append("**⚠ Unconfirmed**\n\n" + "\n\n".join(unconfirmed_lines))
        st.info("**Upcoming viewings (next 7 days)**\n\n" + "\n\n---\n\n".join(parts))


def render_overview(filtered: pd.DataFrame) -> None:
    if filtered.empty:
        st.info("No listings match the current filters.")
        return

    sort_options = {
        "Rent: low to high": ("monthly_rent", True),
        "Rent: high to low": ("monthly_rent", False),
        "MRT: nearest first": ("nearest_mrt_minutes", True),
        "Name: A to Z": ("listing_name", True),
    }

    left, right = st.columns([1.05, 1])
    with left:
        sort_header, sort_control = st.columns([1, 1])
        with sort_header:
            st.subheader("Listings")
        with sort_control:
            sort_choice = st.selectbox("Sort by", list(sort_options.keys()), label_visibility="collapsed")
        sort_col, sort_asc = sort_options[sort_choice]
        filtered = filtered.sort_values(sort_col, ascending=sort_asc, na_position="last")
        display = filtered[
            [
                "source_sheet",
                "listing_name",
                "monthly_rent",
                "nearest_mrt_station",
                "nearest_mrt_minutes",
                "booking_time",
                "booking_confirmed",
                "agent_contact",
                "viewed_unit",
                "viewing_outcome",
                "size_sqft",
                "property_type",
                "review_status",
            ]
        ].rename(
            columns={
                "source_sheet": "List",
                "listing_name": "Listing",
                "monthly_rent": "Rent (S$)",
                "nearest_mrt_station": "Nearest MRT",
                "nearest_mrt_minutes": "Distance to MRT",
                "booking_time": "Booking time",
                "booking_confirmed": "Confirmed",
                "agent_contact": "Agent / Contact",
                "viewed_unit": "Viewed",
                "viewing_outcome": "Outcome",
                "size_sqft": "Size (sqft)",
                "property_type": "Type",
                "review_status": "Checklist",
            }
        )
        st.dataframe(
            display,
            hide_index=True,
            width="stretch",
            column_config={
                "Rent (S$)": st.column_config.NumberColumn(format="S$ %.0f"),
                "Distance to MRT": st.column_config.NumberColumn(format="%.0f min"),
                "Size (sqft)": st.column_config.NumberColumn(format="%.0f"),
            },
        )

    with right:
        st.subheader("Quick cuts")
        property_mix = filtered["property_type"].value_counts().rename_axis("Property type").reset_index(name="Count")
        st.caption("Property mix")
        st.bar_chart(property_mix, x="Property type", y="Count", width="stretch")

        rent_map = filtered.dropna(subset=["size_sqft", "monthly_rent"])
        st.caption("Rent vs size")
        if rent_map.empty:
            st.info("Add more size data in the source sheet to unlock the rent vs size view.")
        else:
            st.scatter_chart(rent_map, x="size_sqft", y="monthly_rent", color="source_sheet")


def render_listing_detail(selected: pd.Series) -> None:
    st.markdown(
        f"""
        <div class="surface-card">
            <h3>{selected['listing_name']}</h3>
            <div>
                <span class="pill">{selected['source_sheet']}</span>
                <span class="pill">{format_currency(selected['monthly_rent'])}</span>
                <span class="pill">{selected['property_type']}</span>
                <span class="pill">{'Co-living' if bool(selected['is_co_living']) else 'Private unit'}</span>
            </div>
            <div class="detail-grid">
                <div class="detail-tile"><div class="label">Address</div><div class="value">{selected['address']}</div></div>
                <div class="detail-tile"><div class="label">Nearest MRT</div><div class="value">{selected['nearest_mrt_station'] or 'Not listed'}</div></div>
                <div class="detail-tile"><div class="label">MRT Distance</div><div class="value">{format_number(selected['nearest_mrt_minutes'], ' min')}</div></div>
                <div class="detail-tile"><div class="label">Booking Time</div><div class="value">{selected['booking_time'] or 'Not listed'}{' ✓' if selected['booking_confirmed'] == 'confirmed' else (' ⚠' if selected['booking_confirmed'] == 'unconfirmed' else '')}</div></div>
                <div class="detail-tile"><div class="label">Agent / Contact</div><div class="value">{selected['agent_contact'] or 'Not listed'}</div></div>
                <div class="detail-tile"><div class="label">Viewed</div><div class="value">{selected['viewed_unit']}</div></div>
                <div class="detail-tile"><div class="label">Outcome</div><div class="value">{selected['viewing_outcome']}</div></div>
                <div class="detail-tile"><div class="label">Beds / Baths</div><div class="value">{selected['beds']} / {selected['baths']}</div></div>
                <div class="detail-tile"><div class="label">Size</div><div class="value">{format_number(selected['size_sqft'], ' sqft')}</div></div>
                <div class="detail-tile"><div class="label">PSF</div><div class="value">{format_number(selected['psf_sgd'])}</div></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    action_columns = st.columns(3)
    with action_columns[0]:
        if selected["listing_url"]:
            st.link_button("Open source listing", selected["listing_url"], width="content")
    with action_columns[1]:
        google_calendar_url = build_google_calendar_url(selected)
        if google_calendar_url is not None:
            st.link_button("Open Google Calendar invite", google_calendar_url, width="content")
    with action_columns[2]:
        invite = build_calendar_invite(selected)
        if invite is not None:
            invite_content, invite_name = invite
            st.download_button(
                "Download calendar invite",
                data=invite_content,
                file_name=invite_name,
                mime="text/calendar",
                width="content",
            )
    st.markdown("**Source notes**")
    st.write(selected["notes"])

    missing_items = missing_checklist_items(selected)
    if missing_items:
        missing_html = "".join(f"<li>{item}</li>" for item in missing_items)
        st.markdown(
            f"""
            <div class="missing-list">
                <strong>Still needs confirmation</strong>
                <ul>{missing_html}</ul>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.progress(float(selected["completion_ratio"]), text=f"Checklist completion: {int(selected['completion_ratio'] * 100)}%")


def render_review_form(selected: pd.Series, reviews: pd.DataFrame) -> None:
    review = review_for_listing(selected["listing_key"], reviews)
    with st.form(f"review-{selected['listing_key']}"):
        st.markdown("**Viewing status**")
        status_left, status_right = st.columns(2)
        with status_left:
            viewed_unit = st.checkbox("Viewed this unit", value=review["viewed_unit"] == "Yes")
        with status_right:
            viewing_outcome = st.selectbox(
                "Viewing outcome",
                VIEWING_OUTCOME_OPTIONS,
                index=option_index(VIEWING_OUTCOME_OPTIONS, review["viewing_outcome"]),
                help="Saved as Pending if the unit has not been viewed yet.",
            )

        st.markdown("**Checklist details**")
        left, right = st.columns(2)
        with left:
            utilities_aircon_inclusive = st.selectbox(
                "1. Is price inclusive of utilities (water, electric) + air con?",
                UTILITIES_OPTIONS,
                index=option_index(UTILITIES_OPTIONS, review["utilities_aircon_inclusive"]),
            )
            water_heater = st.selectbox(
                "2. Water Heater",
                YES_NO_OPTIONS,
                index=option_index(YES_NO_OPTIONS, review["water_heater"]),
            )
            washing_machine = st.selectbox(
                "3. Washing Machine",
                YES_NO_OPTIONS,
                index=option_index(YES_NO_OPTIONS, review["washing_machine"]),
            )
            fridge = st.selectbox(
                "4. Fridge",
                YES_NO_OPTIONS,
                index=option_index(YES_NO_OPTIONS, review["fridge"]),
            )
            bedroom_ceiling_fan = st.selectbox(
                "5. Ceiling Fan in bedroom",
                YES_NO_OPTIONS,
                index=option_index(YES_NO_OPTIONS, review["bedroom_ceiling_fan"]),
            )

        with right:
            living_room_ceiling_fan = st.selectbox(
                "6. Ceiling Fan in living room",
                YES_NO_OPTIONS,
                index=option_index(YES_NO_OPTIONS, review["living_room_ceiling_fan"]),
            )
            floor_level = st.text_input("7. Floor Level of unit", value=review["floor_level"])
            sun_exposure = st.selectbox(
                "8. Unit gets sunrise or sunset?",
                SUN_OPTIONS,
                index=option_index(SUN_OPTIONS, review["sun_exposure"]),
            )
            resident_profile = st.text_area(
                "9. Group of people that stays around the building",
                value=review["resident_profile"],
                height=110,
            )
            co_living_kitchen = st.selectbox(
                "10. Co-Living - kitchen in common area or not?",
                KITCHEN_OPTIONS,
                index=option_index(KITCHEN_OPTIONS, review["co_living_kitchen"]),
                disabled=not bool(selected["is_co_living"]),
                help="Only relevant for co-living listings.",
            )

        overall_notes = st.text_area(
            "General viewing notes",
            value=review["overall_notes"],
            height=120,
        )
        submitted = st.form_submit_button("Save checklist", width="stretch")

    if submitted:
        record = {
            "listing_key": selected["listing_key"],
            "viewed_unit": "Yes" if viewed_unit else "No",
            "viewing_outcome": viewing_outcome if viewed_unit else VIEWING_OUTCOME_OPTIONS[0],
            "utilities_aircon_inclusive": utilities_aircon_inclusive,
            "water_heater": water_heater,
            "washing_machine": washing_machine,
            "fridge": fridge,
            "bedroom_ceiling_fan": bedroom_ceiling_fan,
            "living_room_ceiling_fan": living_room_ceiling_fan,
            "floor_level": floor_level.strip(),
            "sun_exposure": sun_exposure,
            "resident_profile": resident_profile.strip(),
            "co_living_kitchen": co_living_kitchen if bool(selected["is_co_living"]) else "Not applicable",
            "overall_notes": overall_notes.strip(),
            "last_updated": datetime.now().isoformat(timespec="seconds"),
        }
        save_review(record)
        st.success("Checklist saved.")
        st.rerun()


def render_tracker(data: pd.DataFrame) -> None:
    booked = data[data["booking_time"].astype(str).str.strip() != ""].copy()
    if not booked.empty:
        booked["google_calendar_url"] = booked.apply(build_google_calendar_url, axis=1)
        confirmed_booked = booked[booked["booking_confirmed"] == "confirmed"]
        unconfirmed_booked = booked[booked["booking_confirmed"] != "confirmed"]

        def _render_booked_table(df: pd.DataFrame) -> None:
            view = df[
                [
                    "listing_name",
                    "booking_time",
                    "agent_contact",
                    "address",
                    "google_calendar_url",
                ]
            ].rename(
                columns={
                    "listing_name": "Listing",
                    "booking_time": "Booking time",
                    "agent_contact": "Agent / Contact",
                    "address": "Address",
                    "google_calendar_url": "Google Calendar",
                }
            )
            st.dataframe(
                view,
                hide_index=True,
                width="stretch",
                column_config={
                    "Google Calendar": st.column_config.LinkColumn(
                        "Google Calendar",
                        display_text="Open invite",
                    )
                },
            )

        if not confirmed_booked.empty:
            st.subheader("✓ Confirmed Viewings")
            _render_booked_table(confirmed_booked)

        if not unconfirmed_booked.empty:
            st.subheader("⚠ Unconfirmed Viewings")
            _render_booked_table(unconfirmed_booked)

    st.subheader("Review Tracker")
    reviewed = data[data["review_status"] != "Not reviewed"].copy()
    st.caption(f"Saved reviews are written to {REVIEWS_PATH.name} in the project root.")
    if reviewed.empty:
        st.info("No checklist entries saved yet.")
        return

    tracker = reviewed[
        [
            "listing_name",
            "source_sheet",
            "viewed_unit",
            "viewing_outcome",
            "review_status",
            "answered_fields",
            "total_fields",
            "utilities_aircon_inclusive",
            "water_heater",
            "washing_machine",
            "fridge",
            "sun_exposure",
            "co_living_kitchen",
            "last_updated",
        ]
    ].rename(
        columns={
            "listing_name": "Listing",
            "source_sheet": "List",
            "viewed_unit": "Viewed",
            "viewing_outcome": "Outcome",
            "review_status": "Status",
            "answered_fields": "Answered",
            "total_fields": "Required",
            "utilities_aircon_inclusive": "Utilities + AC",
            "water_heater": "Water heater",
            "washing_machine": "Washing machine",
            "fridge": "Fridge",
            "sun_exposure": "Sun",
            "co_living_kitchen": "Kitchen",
            "last_updated": "Updated",
        }
    )
    st.dataframe(tracker, hide_index=True, width="stretch")
    if REVIEWS_PATH.exists():
        st.download_button(
            "Download saved checklist CSV",
            data=Path(REVIEWS_PATH).read_bytes(),
            file_name=REVIEWS_PATH.name,
            mime="text/csv",
            width="stretch",
        )


def main() -> None:
    st.set_page_config(page_title="Unit Viewing Dashboard", layout="wide")
    inject_styles()

    if WORKBOOK_PATH.exists():
        st.caption(f"Using local workbook: {WORKBOOK_PATH.name}")
        listings = load_listings(str(WORKBOOK_PATH), workbook_signature(WORKBOOK_PATH))
    else:
        st.info("Upload your Excel workbook to use the dashboard. The repository does not store the source workbook.")
        uploaded_workbook = st.file_uploader(
            "Upload property_listings_MM.xlsx",
            type=["xlsx"],
            accept_multiple_files=False,
        )
        if uploaded_workbook is None:
            st.stop()
        listings = load_uploaded_listings(uploaded_workbook.getvalue(), uploaded_workbook.name)

    reviews = load_reviews(str(REVIEWS_PATH))
    dashboard = build_dashboard_data(listings, reviews)
    filtered = filter_dashboard(dashboard)

    render_hero(filtered)

    if filtered.empty:
        st.info("No listings match the current filters.")
        return

    overview_tab, review_tab, tracker_tab = st.tabs(["Overview", "Unit Review", "Tracker"])

    with overview_tab:
        render_overview(filtered)

    with review_tab:
        labels = {
            row["listing_key"]: listing_label(row)
            for _, row in filtered.iterrows()
        }
        listing_keys = filtered["listing_key"].tolist()
        previous_key = st.session_state.get("selected_listing_key")
        default_index = listing_keys.index(previous_key) if previous_key in listing_keys else 0
        selected_key = st.selectbox(
            "Choose a listing to review",
            listing_keys,
            index=default_index,
            format_func=lambda key: labels[key],
            key="selected_listing_key",
        )
        selected = filtered.loc[filtered["listing_key"] == selected_key].iloc[0]
        left, right = st.columns([1.15, 1])
        with left:
            render_listing_detail(selected)
        with right:
            st.subheader("Viewing checklist & outcome")
            render_review_form(selected, reviews)

    with tracker_tab:
        render_tracker(dashboard)


if __name__ == "__main__":
    main()